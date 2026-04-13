import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from copy import copy  # Standard python copy for styles

def final_sync_with_injection(source_file, target_file, output_file):
    print(f"📂 Loading Workbooks...")
    wb_source = openpyxl.load_workbook(source_file)
    wb_target = openpyxl.load_workbook(target_file)
    ws = wb_target.active 
    
    # Ensure Sheet2 exists in target for lookup data
    if 'Sheet2' not in wb_target.sheetnames:
        print("⚠️ Sheet2 not found in target. Copying from source...")
        source_s2 = wb_source['Sheet2']
        target_s2 = wb_target.create_sheet('Sheet2')
        for row in source_s2.rows:
            for cell in row:
                target_s2[cell.coordinate].value = cell.value

    ws2 = wb_target['Sheet2']

    # --- 1. Transfer Named Ranges ---
    print("🔗 Syncing Named Ranges...")
    for name, defn in wb_source.defined_names.items():
        try:
            new_defn = DefinedName(name=defn.name, attr_text=defn.value)
            wb_target.defined_names[name] = new_defn
        except Exception as e:
            print(f"  ⚠️ Skip name '{name}': {e}")

    # --- 2. Inject New Columns next to "City, County, State" ---
    # We iterate backwards to maintain column index integrity during insertion
    header_cells = list(ws[1])
    for i in range(len(header_cells) - 1, -1, -1):
        cell = header_cells[i]
        if str(cell.value).strip() == "City, County, State":
            col_to_right = cell.column + 1
            print(f"📍 Injecting columns at index {col_to_right}...")
            ws.insert_cols(col_to_right, 3)
            
            new_headers = ["State", "County", "Coordinates"]
            for offset, name in enumerate(new_headers):
                new_cell = ws.cell(row=1, column=col_to_right + offset)
                new_cell.value = name
                # Copy formatting from the source column
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.alignment = copy(cell.alignment)

    # --- 3. Map All Matching Column Indices (Including Injected Ones) ---
    header = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    
    state_indices = [i+1 for i, name in enumerate(header) if name == "State"]
    county_indices = [i+1 for i, name in enumerate(header) if name == "County"]
    coord_indices = [i+1 for i, name in enumerate(header) if name == "Coordinates"]

    print(f"🔍 Logic Target: {len(state_indices)} State(s), {len(county_indices)} County(ies).")

    # --- 4. Setup State Validation ---
    # Points to Sheet2 Column Z (the state list)
    state_dv = DataValidation(type="list", formula1="Sheet2!$Z$1:$Z$51", allow_blank=True)
    ws.add_data_validation(state_dv)

    # --- 5. Apply Logic Row by Row ---
    print("⚙️ Applying Logic to all identified columns...")
    for r in range(2, ws.max_row + 1):
        
        # A. Apply State Dropdowns
        for s_idx in state_indices:
            state_dv.add(ws.cell(row=r, column=s_idx))

        # B. Apply County Dependent Dropdowns
        for c_idx in county_indices:
            left_states = [s for s in state_indices if s < c_idx]
            ref_state = left_states[-1] if left_states else state_indices[0]
            
            county_formula = f'=INDIRECT("S_"&INDIRECT(ADDRESS(ROW(),{ref_state})))'
            county_dv = DataValidation(type="list", formula1=county_formula, allow_blank=True)
            ws.add_data_validation(county_dv)
            county_dv.add(ws.cell(row=r, column=c_idx))

        # C. Apply Coordinate VLOOKUPs
        for cr_idx in coord_indices:
            left_states = [s for s in state_indices if s < cr_idx]
            left_counties = [c for c in county_indices if c < cr_idx]
            
            ref_s_idx = left_states[-1] if left_states else state_indices[0]
            ref_c_idx = left_counties[-1] if left_counties else county_indices[0]
            
            s_cell = ws.cell(row=r, column=ref_s_idx).coordinate
            c_cell = ws.cell(row=r, column=ref_c_idx).coordinate
            
            # VLOOKUP pulls from Sheet2 Columns E (Key) and F (Coords)
            lookup_formula = f'=IFERROR(VLOOKUP({s_cell}&"-"&{c_cell}, Sheet2!$E:$F, 2, FALSE), "")'
            ws.cell(row=r, column=cr_idx).value = lookup_formula

    wb_target.save(output_file)
    print(f"\n✨ Success! Columns injected and logic applied. Saved as: {output_file}")

if __name__ == "__main__":
    SRC = 'Ancestors Database_Osama_v02_modified.xlsx'
    TGT = 'Ancestors Database_v01_copy.xlsx'
    FINAL = 'Ancestors_Database_v05.xlsx'
    
    final_sync_with_injection(SRC, TGT, FINAL)