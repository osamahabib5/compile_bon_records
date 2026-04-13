import pandas as pd
import spacy
import re
import time
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut

# Load spaCy
try:
    nlp = spacy.load("en_core_web_md")
except OSError:
    from spacy.cli import download
    download("en_core_web_md")
    nlp = spacy.load("en_core_web_md")

geolocator = Nominatim(user_agent="historical_research_v7_update")

# --- 1. Load Files ---
file_a_path = 'Consolidated_Directory.xlsx'
file_b_path = 'Book_of_Negroes_Copy.xlsx'
file_c_path = 'US_Counties_Coordinates.xlsx' 

df_a = pd.read_excel(file_a_path)
df_b = pd.read_excel(file_b_path)
df_c = pd.read_excel(file_c_path)

df_b.columns = [col.replace(" ", "_") for col in df_b.columns]

# List of known countries/major regions to validate against
VALID_COUNTRIES = {"United Kingdom", "Jamaica", "Madagascar", "Africa", "England", "France", "Spain", "Canada"}

# --- 2. Processing & Extraction Logic ---

def get_coordinates_cascading(city, county, state, country, df_lookup):
    """
    Coordinates logic: 
    1. Check Sheet2 (df_lookup) if it's a US location.
    2. Use Geopy if any geo field is present.
    3. Return '-' otherwise.
    """
    # Attempt 1: Sheet2 Lookup (US Only)
    if country == "United States":
        match = df_lookup[(df_lookup['State'] == state) & (df_lookup['County'] == county)]
        if not match.empty:
            return f"{match.iloc[0]['Latitude']}, {match.iloc[0]['Longitude']}"

    # Attempt 2: Geopy API (Fallback for US or Primary for International)
    # Only run if at least one field is populated
    geo_parts = [p for p in [city, county, state, country] if p and p != "-"]
    if geo_parts:
        query = ", ".join(geo_parts)
        try:
            time.sleep(1)  # Respect API usage limits
            location = geolocator.geocode(query, timeout=10)
            if location:
                return f"{location.latitude}, {location.longitude}"
        except:
            pass
            
    return "-"

def process_entry(row):
    notes = str(row['Notes']) if pd.notna(row['Notes']) else ""
    
    # 1. Extract Enslaver
    enslaver = "-"
    enslaver_match = re.search(r'\((.*?)\)', notes)
    if enslaver_match:
        content = enslaver_match.group(1).strip()
        enslaver = "-" if "own bottom" in content.lower() else content
    
    # Clean text for NLP
    clean_notes = re.sub(r'\(.*?\)', '', notes).strip()
    doc = nlp(clean_notes)
    
    # 2. Identify Geography
    city, county, state, country = "-", "-", "-", "United States"
    
    # Keyword Search for County (e.g., "Princess Ann County")
    # Matches words preceding the word 'County'
    county_kw_match = re.search(r'([A-Z][a-z]+(?:\s[A-Z][a-z]+)?)\s+County', clean_notes)
    # Matches "Parish of..."
    parish_match = re.search(r'parish of\s+([A-Z][a-z]+(?:\s[A-Z][a-z]+)?)', clean_notes, re.I)

    gpes = [ent.text for ent in doc.ents if ent.label_ in ["GPE", "LOC"]]
    
    # Country Validation
    found_country = False
    for gpe in gpes:
        if gpe in VALID_COUNTRIES:
            country = gpe
            found_country = True
            break
    
    # Assign locations based on GPE extraction
    if len(gpes) >= 3: city, county, state = gpes[0], gpes[1], gpes[2]
    elif len(gpes) == 2: city, state = gpes[0], gpes[1]
    elif len(gpes) == 1: state = gpes[0]
    
    # Override County if keyword found (Priority)
    if county_kw_match:
        county = county_kw_match.group(1).strip()
    elif parish_match:
        county = parish_match.group(1).strip()

    # Final logic: if state is a known US state, country MUST be US
    us_states = set(df_c['State'].unique())
    if state in us_states or (county != "-" and county in set(df_c['County'].unique())):
        country = "United States"

    # 3. Get Coordinates
    coords = get_coordinates_cascading(city, county, state, country, df_c)

    return pd.Series([enslaver, city, county, state, country, coords])

# Apply extraction
print("Processing records and fetching coordinates...")
df_a[['Enslaver', 'Extracted_City', 'Extracted_County', 'Extracted_State', 'Country', 'Departure_Coordinates']] = df_a.apply(process_entry, axis=1)

# Clean up columns
if 'Description' in df_a.columns:
    df_a.rename(columns={'Description': 'Origin'}, inplace=True)

if 'Age' in df_a.columns:
    df_a['Birthdate'] = df_a['Age'].apply(lambda x: 1783 - int(float(x)) if pd.notna(x) and str(x).replace('.','').isdigit() else "-")
    df_a.drop(columns=['Age'], inplace=True)

df_a['State'] = ""
df_a['County'] = ""

# --- 3. Merge with File B ---
df_a['Match_Name'] = (df_a['First_Name'].fillna('').astype(str).str.strip() + " " + 
                      df_a['Surname'].fillna('').astype(str).str.strip()).str.lower()
df_a['Match_Ship'] = df_a['Ship_Name'].fillna('').astype(str).str.strip().str.lower()

if 'Name' in df_b.columns:
    df_b['Match_Name'] = df_b['Name'].fillna('').astype(str).str.strip().str.lower()
if 'Ship_Name' in df_b.columns:
    df_b['Match_Ship'] = df_b['Ship_Name'].fillna('').astype(str).str.strip().str.lower()

target_cols = ['Ref_Page', 'Origination_Port', 'Departure_Port', 'Departure_Date', 'Arrival_Port_City', 'Ship_Notes']
lookup_cols = ['Match_Ship', 'Match_Name'] + [c for c in target_cols if c in df_b.columns]

df_b_match = df_b[lookup_cols].drop_duplicates(subset=['Match_Ship', 'Match_Name'])
df_final = df_a.merge(df_b_match, on=['Match_Ship', 'Match_Name'], how='left', suffixes=('', '_LKP'))

# --- 4. Reorganize Columns ---
fixed_start = ['Ship_Name', 'Notes', 'Ship_Notes', 'First_Name', 'Surname', 'Birthdate']
geo_block = ['Origin', 'Extracted_City', 'Extracted_County', 'Extracted_State', 'Country', 'State', 'County', 'Departure_Coordinates']

all_current_cols = df_final.columns.tolist()
remaining = [c for c in all_current_cols if c not in fixed_start + geo_block and not c.endswith('_LKP') and 'Match_' not in c]

df_final = df_final[fixed_start + remaining + geo_block]

# --- 5. Export & Dynamic Excel UI ---
output_file = 'Consolidated_Directory_v6.xlsx'
df_c = df_c.sort_values(['State', 'County'])

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df_final.to_excel(writer, index=False, sheet_name='Main')
    df_c.to_excel(writer, index=False, sheet_name='Sheet2')

wb = load_workbook(output_file)
ws_main = wb['Main']
header = {cell.value: get_column_letter(cell.column) for cell in ws_main[1]}

# Prepare Dropdowns
unique_states = sorted(df_c['State'].unique().astype(str).tolist())
for i, s in enumerate(unique_states, 1):
    wb['Sheet2'].cell(row=i, column=10).value = s 

dv_state = DataValidation(type="list", formula1=f"Sheet2!$J$1:$J${len(unique_states)}", allow_blank=True)
ws_main.add_data_validation(dv_state)

for r in range(2, ws_main.max_row + 1):
    # State dropdown
    dv_state.add(ws_main[f"{header['State']}{r}"])
    
    # Dependent County dropdown
    st_cell = f"{header['State']}{r}"
    formula_co = f'OFFSET(Sheet2!$A$1, MATCH({st_cell}, Sheet2!$B:$B, 0)-1, 0, COUNTIF(Sheet2!$B:$B, {st_cell}), 1)'
    dv_county = DataValidation(type="list", formula1=formula_co, allow_blank=True)
    ws_main.add_data_validation(dv_county)
    dv_county.add(ws_main[f"{header['County']}{r}"])

wb.save(output_file)
print(f"✨ Success! Script finished. File saved as: {output_file}")