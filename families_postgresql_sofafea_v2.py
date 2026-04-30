import os
import re
from datetime import date, datetime

import pandas as pd
import psycopg2
from dotenv import load_dotenv
from geopy.extra.rate_limiter import RateLimiter
from geopy.geocoders import Nominatim
from openpyxl import load_workbook


load_dotenv()


DB_CONNECTION_STRING = os.getenv("DB_CONNECTION_STRING")
INPUT_FILE = "Database_template_records_insertion_JO_Attaquin_copy.xlsx"
OUTPUT_FILE = "Database_template_records_insertion_JO_Attaquin_v1.xlsx"
DEFAULT_COUNTRY = "United States"
DEFAULT_LANDMARK = "-"
LOCATION_HEADER = "City, Count, State"
COORDINATES_HEADER = "Coordinates"


geolocator = Nominatim(user_agent="genealogy_family_ingestor_v2", timeout=10)
geocode_service = RateLimiter(
    geolocator.geocode,
    min_delay_seconds=1.2,
    max_retries=2,
    error_wait_seconds=5.0,
    swallow_exceptions=True,
)


def clean_val(val):
    if val is None:
        return None

    if isinstance(val, float) and pd.isna(val):
        return None

    if pd.isna(val):
        return None

    cleaned = str(val).strip()
    if cleaned == "" or cleaned.lower() in {"nan", "none"} or cleaned == "-":
        return None

    return cleaned


def format_date(val):
    cleaned = clean_val(val)
    if not cleaned:
        return None

    if isinstance(val, pd.Timestamp):
        return val.date().isoformat()

    if isinstance(val, datetime):
        return val.date().isoformat()

    if isinstance(val, date):
        return val.isoformat()

    try:
        numeric = float(cleaned)
        if numeric.is_integer() and len(str(int(numeric))) == 4:
            return f"{int(numeric)}-01-01"
    except Exception:
        pass

    if re.fullmatch(r"\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}", cleaned):
        return cleaned[:10]

    parsed = pd.to_datetime(cleaned, errors="coerce", dayfirst=True)
    if pd.notna(parsed):
        return parsed.date().isoformat()

    return None


def normalize_header(val):
    cleaned = clean_val(val)
    return cleaned if cleaned else ""


def normalize_place_key(place):
    cleaned = clean_val(place)
    if not cleaned:
        return None

    return re.sub(r"\s+", " ", cleaned).strip().lower()


def parse_generation_number(text, fallback=None):
    cleaned = clean_val(text)
    if not cleaned:
        return fallback

    match = re.search(r"generation\s*(\d+)", cleaned, re.IGNORECASE)
    if match:
        return int(match.group(1))

    return fallback


def find_generation_marker(row):
    for value in row.tolist():
        cleaned = clean_val(value)
        if cleaned and cleaned.lower().startswith("generation"):
            return cleaned
    return None


def row_has_any_data(row):
    return any(clean_val(value) for value in row.tolist())


def infer_spouse_gender(subject_gender):
    gender = clean_val(subject_gender)
    if not gender:
        return None

    lowered = gender.lower()
    if lowered == "male":
        return "Female"
    if lowered == "female":
        return "Male"
    return None


def get_db_connection():
    if not DB_CONNECTION_STRING:
        print("DB_CONNECTION_STRING not found in the environment.")
        return None

    try:
        return psycopg2.connect(DB_CONNECTION_STRING)
    except Exception as exc:
        print(f"Failed to connect to PostgreSQL: {exc}")
        return None


def ensure_coordinate_columns(ws):
    for col_idx in range(ws.max_column, 0, -1):
        header = normalize_header(ws.cell(row=1, column=col_idx).value)
        if header != LOCATION_HEADER:
            continue

        next_header = normalize_header(ws.cell(row=1, column=col_idx + 1).value)
        if next_header == COORDINATES_HEADER:
            continue

        ws.insert_cols(col_idx + 1)
        ws.cell(row=1, column=col_idx + 1).value = COORDINATES_HEADER


def get_location_column_pairs(ws):
    pairs = []
    for col_idx in range(1, ws.max_column + 1):
        header = normalize_header(ws.cell(row=1, column=col_idx).value)
        if header == LOCATION_HEADER:
            next_header = normalize_header(ws.cell(row=1, column=col_idx + 1).value)
            if next_header == COORDINATES_HEADER:
                pairs.append((col_idx, col_idx + 1))
    return pairs


def geocode_place(place):
    cleaned = clean_val(place)
    if not cleaned:
        return None

    queries = [cleaned]
    if DEFAULT_COUNTRY.lower() not in cleaned.lower() and "usa" not in cleaned.lower():
        queries.insert(0, f"{cleaned}, {DEFAULT_COUNTRY}")

    for query in queries:
        try:
            location = geocode_service(query)
        except Exception as exc:
            print(f"Geocoding error for '{query}': {exc}")
            location = None

        if location:
            return f"{location.latitude}, {location.longitude}"

    return None


def enrich_workbook_with_coordinates(input_path, output_path):
    if not os.path.exists(input_path):
        print(f"File {input_path} not found.")
        return None

    workbook = load_workbook(input_path)

    for worksheet in workbook.worksheets:
        ensure_coordinate_columns(worksheet)

    coord_cache = {}
    places_to_geocode = {}

    for worksheet in workbook.worksheets:
        location_pairs = get_location_column_pairs(worksheet)
        for row_idx in range(2, worksheet.max_row + 1):
            for place_col, coord_col in location_pairs:
                place = clean_val(worksheet.cell(row=row_idx, column=place_col).value)
                coord = clean_val(worksheet.cell(row=row_idx, column=coord_col).value)
                key = normalize_place_key(place)

                if not key:
                    continue

                if coord:
                    coord_cache[key] = coord
                else:
                    places_to_geocode[key] = place

    missing_places = [place for key, place in places_to_geocode.items() if key not in coord_cache]
    total_missing = len(missing_places)

    if total_missing:
        print(f"Geocoding {total_missing} unique place values across all sheets...")

    for idx, place in enumerate(missing_places, start=1):
        coord_cache[normalize_place_key(place)] = geocode_place(place)
        if idx == total_missing or idx % 10 == 0:
            print(f"Geocoded {idx}/{total_missing} unique locations")

    for worksheet in workbook.worksheets:
        location_pairs = get_location_column_pairs(worksheet)
        for row_idx in range(2, worksheet.max_row + 1):
            for place_col, coord_col in location_pairs:
                existing_coord = clean_val(worksheet.cell(row=row_idx, column=coord_col).value)
                if existing_coord:
                    continue

                place = clean_val(worksheet.cell(row=row_idx, column=place_col).value)
                place_key = normalize_place_key(place)
                if not place_key:
                    continue

                resolved_coord = coord_cache.get(place_key)
                if resolved_coord:
                    worksheet.cell(row=row_idx, column=coord_col).value = resolved_coord

    workbook.save(output_path)
    print(f"Workbook saved with coordinate columns to: {output_path}")
    return output_path


def get_or_insert_location(
    cur,
    city,
    county,
    state,
    coords,
    country=DEFAULT_COUNTRY,
    landmark=DEFAULT_LANDMARK,
):
    city = clean_val(city)
    county = clean_val(county)
    state = clean_val(state)
    coords = clean_val(coords)
    country = clean_val(country) or DEFAULT_COUNTRY
    landmark = clean_val(landmark) or DEFAULT_LANDMARK

    if not any([city, county, state, coords]):
        return None

    cur.execute(
        """
        SELECT location_id
        FROM locations
        WHERE city IS NOT DISTINCT FROM %s
          AND county IS NOT DISTINCT FROM %s
          AND state IS NOT DISTINCT FROM %s
          AND country IS NOT DISTINCT FROM %s
          AND landmark IS NOT DISTINCT FROM %s
          AND coordinates IS NOT DISTINCT FROM %s
        """,
        (city, county, state, country, landmark, coords),
    )

    existing = cur.fetchone()
    if existing:
        return existing[0]

    cur.execute(
        """
        INSERT INTO locations (city, county, state, country, landmark, coordinates)
        VALUES (%s, %s, %s, %s, %s, %s)
        RETURNING location_id
        """,
        (city, county, state, country, landmark, coords),
    )
    return cur.fetchone()[0]


def update_member_if_missing(cur, member_id, fields):
    cur.execute(
        """
        UPDATE family_members
        SET alias = COALESCE(alias, %s),
            gender = COALESCE(gender, %s),
            race = COALESCE(race, %s),
            ethnicity = COALESCE(ethnicity, %s),
            father_id = COALESCE(father_id, %s),
            mother_id = COALESCE(mother_id, %s),
            birth_date = COALESCE(birth_date, %s),
            birth_location_id = COALESCE(birth_location_id, %s),
            death_date = COALESCE(death_date, %s),
            death_location_id = COALESCE(death_location_id, %s),
            marriage_date = COALESCE(marriage_date, %s),
            marriage_location_id = COALESCE(marriage_location_id, %s)
        WHERE member_id = %s
        """,
        (
            fields["alias"],
            fields["gender"],
            fields["race"],
            fields["ethnicity"],
            fields["father_id"],
            fields["mother_id"],
            fields["birth_date"],
            fields["birth_location_id"],
            fields["death_date"],
            fields["death_location_id"],
            fields["marriage_date"],
            fields["marriage_location_id"],
            member_id,
        ),
    )


def get_or_insert_member(
    cur,
    first_name,
    last_name,
    generation_number,
    alias=None,
    gender=None,
    race=None,
    ethnicity=None,
    father_id=None,
    mother_id=None,
    birth_date=None,
    birth_location_id=None,
    death_date=None,
    death_location_id=None,
    marriage_date=None,
    marriage_location_id=None,
):
    first_name = clean_val(first_name)
    last_name = clean_val(last_name)

    if not first_name and not last_name:
        return None

    member_fields = {
        "alias": clean_val(alias),
        "gender": clean_val(gender),
        "race": clean_val(race),
        "ethnicity": clean_val(ethnicity),
        "father_id": father_id,
        "mother_id": mother_id,
        "birth_date": format_date(birth_date),
        "birth_location_id": birth_location_id,
        "death_date": format_date(death_date),
        "death_location_id": death_location_id,
        "marriage_date": format_date(marriage_date),
        "marriage_location_id": marriage_location_id,
    }

    cur.execute(
        """
        SELECT member_id
        FROM family_members
        WHERE LOWER(first_name) IS NOT DISTINCT FROM LOWER(%s)
          AND LOWER(last_name) IS NOT DISTINCT FROM LOWER(%s)
          AND generation_number = %s
        """,
        (first_name, last_name, generation_number),
    )
    existing = cur.fetchone()
    if existing:
        update_member_if_missing(cur, existing[0], member_fields)
        return existing[0]

    cur.execute(
        """
        INSERT INTO family_members (
            first_name,
            last_name,
            alias,
            generation_number,
            gender,
            race,
            ethnicity,
            father_id,
            mother_id,
            birth_date,
            birth_location_id,
            death_date,
            death_location_id,
            marriage_date,
            marriage_location_id
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        RETURNING member_id
        """,
        (
            first_name,
            last_name,
            member_fields["alias"],
            generation_number,
            member_fields["gender"],
            member_fields["race"],
            member_fields["ethnicity"],
            member_fields["father_id"],
            member_fields["mother_id"],
            member_fields["birth_date"],
            member_fields["birth_location_id"],
            member_fields["death_date"],
            member_fields["death_location_id"],
            member_fields["marriage_date"],
            member_fields["marriage_location_id"],
        ),
    )
    return cur.fetchone()[0]


def link_spouses(cur, left_member_id, right_member_id):
    if not left_member_id or not right_member_id or left_member_id == right_member_id:
        return

    cur.execute(
        "UPDATE family_members SET spouse_id = %s WHERE member_id = %s",
        (right_member_id, left_member_id),
    )
    cur.execute(
        "UPDATE family_members SET spouse_id = %s WHERE member_id = %s",
        (left_member_id, right_member_id),
    )


def get_subject_locations(row):
    birth_location_id = {
        "city": row.get("City, Count, State"),
        "coords": row.get("Coordinates"),
        "state": row.get("State"),
        "county": row.get("County"),
    }
    death_location_id = {
        "city": row.get("City, Count, State.1"),
        "coords": row.get("Coordinates.1"),
        "state": row.get("State.1"),
        "county": row.get("County.1"),
    }
    marriage_location_id = {
        "city": row.get("City, Count, State.2"),
        "coords": row.get("Coordinates.2"),
        "state": row.get("State.2"),
        "county": row.get("County.2"),
    }
    return birth_location_id, death_location_id, marriage_location_id


def build_location_id(cur, location_payload):
    return get_or_insert_location(
        cur,
        location_payload["city"],
        location_payload["county"],
        location_payload["state"],
        location_payload["coords"],
    )


def process_spouse_record(cur, row, generation_number, subject_id, subject_gender):
    spouse_first_name = row.get("Spouse FirstName")
    spouse_last_name = row.get("Surname.1")

    if not clean_val(spouse_first_name) and not clean_val(spouse_last_name):
        return None

    parent_generation = generation_number - 1

    spouse_father_death_loc_id = get_or_insert_location(
        cur,
        row.get("City"),
        row.get("County.4"),
        row.get("State.4"),
        None,
    )
    spouse_mother_death_loc_id = get_or_insert_location(
        cur,
        row.get("City.1"),
        row.get("County.5"),
        row.get("State.5"),
        None,
    )

    spouse_father_id = get_or_insert_member(
        cur,
        row.get("Father FirstName.1"),
        row.get("Father Surname.1"),
        parent_generation,
        gender="Male",
        race=row.get("Race.2"),
        ethnicity=row.get("Ethnicity.2"),
        death_date=row.get("Death Date.1"),
        death_location_id=spouse_father_death_loc_id,
    )
    spouse_mother_id = get_or_insert_member(
        cur,
        row.get("Mother FirstName.1"),
        row.get("Mother Surname.1"),
        parent_generation,
        gender="Female",
        race=row.get("Race.3"),
        ethnicity=row.get("Ethnicity.3"),
        death_date=row.get("Death_Date"),
        death_location_id=spouse_mother_death_loc_id,
    )
    link_spouses(cur, spouse_father_id, spouse_mother_id)

    spouse_birth_loc_id = get_or_insert_location(
        cur,
        row.get("City, Count, State.3"),
        row.get("County.3"),
        row.get("State.3"),
        row.get("Coordinates.3"),
    )
    spouse_marriage_loc_id = get_or_insert_location(
        cur,
        row.get("City, Count, State.2"),
        row.get("County.2"),
        row.get("State.2"),
        row.get("Coordinates.2"),
    )

    spouse_id = get_or_insert_member(
        cur,
        spouse_first_name,
        spouse_last_name,
        generation_number,
        alias=row.get("Alias.1"),
        gender=infer_spouse_gender(subject_gender),
        race=row.get("Race.1"),
        ethnicity=row.get("Ethnicity.1"),
        father_id=spouse_father_id,
        mother_id=spouse_mother_id,
        birth_date=row.get("Birthdate.1"),
        birth_location_id=spouse_birth_loc_id,
        marriage_date=row.get("Marriage_Date 2Spouse"),
        marriage_location_id=spouse_marriage_loc_id,
    )
    link_spouses(cur, subject_id, spouse_id)
    return spouse_id


def process_sheet(cur, sheet_name, df):
    current_generation = None
    last_subject_id = None
    last_subject_gender = None

    for row_index, row in df.iterrows():
        excel_row_num = row_index + 2
        generation_marker = find_generation_marker(row)

        if generation_marker:
            current_generation = parse_generation_number(generation_marker, current_generation)
            last_subject_id = None
            last_subject_gender = None
            print(f"[{sheet_name}] Switched to Generation {current_generation} at row {excel_row_num}")
            continue

        if not row_has_any_data(row):
            continue

        if current_generation is None:
            print(f"[{sheet_name}] Skipping row {excel_row_num}: no generation marker found yet.")
            continue

        first_name = clean_val(row.get("First Name"))
        last_name = clean_val(row.get("Surname"))
        spouse_first_name = clean_val(row.get("Spouse FirstName"))
        spouse_last_name = clean_val(row.get("Surname.1"))

        if not first_name and not last_name:
            if last_subject_id and (spouse_first_name or spouse_last_name):
                process_spouse_record(cur, row, current_generation, last_subject_id, last_subject_gender)
                print(f"[{sheet_name}] Added continuation spouse row {excel_row_num}")
            continue

        parent_generation = current_generation - 1

        subject_father_id = get_or_insert_member(
            cur,
            row.get("Father FirstName"),
            row.get("Father Surname"),
            parent_generation,
            gender="Male",
        )
        subject_mother_id = get_or_insert_member(
            cur,
            row.get("Mother FirstName"),
            row.get("Mother Surname"),
            parent_generation,
            gender="Female",
        )
        link_spouses(cur, subject_father_id, subject_mother_id)

        birth_loc_payload, death_loc_payload, marriage_loc_payload = get_subject_locations(row)
        subject_birth_loc_id = build_location_id(cur, birth_loc_payload)
        subject_death_loc_id = build_location_id(cur, death_loc_payload)
        subject_marriage_loc_id = build_location_id(cur, marriage_loc_payload)

        subject_id = get_or_insert_member(
            cur,
            row.get("First Name"),
            row.get("Surname"),
            current_generation,
            alias=row.get("Alias"),
            gender=row.get("Gender"),
            race=row.get("Race"),
            ethnicity=row.get("Ethnicity"),
            father_id=subject_father_id,
            mother_id=subject_mother_id,
            birth_date=row.get("Birthdate"),
            birth_location_id=subject_birth_loc_id,
            death_date=row.get("Death Date"),
            death_location_id=subject_death_loc_id,
            marriage_date=row.get("Marriage_Date 2Spouse"),
            marriage_location_id=subject_marriage_loc_id,
        )

        if spouse_first_name or spouse_last_name:
            process_spouse_record(cur, row, current_generation, subject_id, row.get("Gender"))

        last_subject_id = subject_id
        last_subject_gender = row.get("Gender")
        print(
            f"[{sheet_name}] Processed row {excel_row_num}: "
            f"{clean_val(row.get('First Name'))} {clean_val(row.get('Surname'))} "
            f"(Gen {current_generation})"
        )


def run_genealogy_ingestion(file_path):
    if not os.path.exists(file_path):
        print(f"File {file_path} not found.")
        return

    workbook_frames = pd.read_excel(file_path, sheet_name=None)

    conn = get_db_connection()
    if not conn:
        return

    cur = conn.cursor()

    try:
        for sheet_name, dataframe in workbook_frames.items():
            dataframe.columns = dataframe.columns.astype(str).str.strip()
            process_sheet(cur, sheet_name, dataframe)

        conn.commit()
        print("Ingestion complete.")
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()


def main():
    enriched_file = enrich_workbook_with_coordinates(INPUT_FILE, OUTPUT_FILE)
    if not enriched_file:
        return

    run_genealogy_ingestion(enriched_file)


if __name__ == "__main__":
    main()
