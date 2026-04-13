import os
import pandas as pd
import re
import time
from google import genai
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from geopy.geocoders import Nominatim


# --- 1. Configuration & Rate Limiting ---
api_key = os.getenv("GEMINI_API_KEY")
if not api_key:
    raise ValueError("GEMINI_API_KEY not found. Ensure your .env file is configured.")

client = genai.Client(api_key=api_key)
geolocator = Nominatim(user_agent="historical_research_v10_final")

# 15 RPM limit requires ~4.1 seconds between requests to be safe
REQUEST_DELAY = 4.1 

def call_gemini_ner(notes_text):
    """
    Uses Gemini to analyze the 'Notes' column with rate limiting 
    and exponential backoff for 503/server errors.
    """
    if not notes_text or pd.isna(notes_text) or str(notes_text).strip() == "":
        return ["-", "-", "-", "-", "United States"]

    prompt = f"""
    Refer to the 'Notes' column of this historical record to extract the correct data for:
    Enslaver | Extracted_City | Extracted_County | Extracted_State | Country

    Rules:
    1. Enslaver: Name inside parentheses. If it says 'on his own bottom' or similar, return '-'.
    2. Extracted_County: Specifically look for keywords like "Princess Ann County" or "Parish of". 
    3. Geography: Identify birthplace/origin. If a US State is mentioned, Country is 'United States'.
    4. Format: Return ONLY the values separated by pipes (|).

    Notes Column: "{notes_text}"
    """
    
    max_retries = 5
    base_backoff = 5  # Seconds
    
    for attempt in range(max_retries):
        try:
            # Respect the 15 RPM limit on every attempt
            time.sleep(REQUEST_DELAY) 
            
            response = client.models.generate_content(
                model="gemini-3.1-flash-lite-preview", 
                contents=prompt
            )
            
            parts = [p.strip() for p in response.text.strip().split('|')]
            if len(parts) == 5:
                return parts
            break 
            
        except Exception as e:
            error_str = str(e)
            # Handle Rate Limits (429) and Server Overload (503)
            if "429" in error_str or "503" in error_str:
                wait_time = base_backoff * (2 ** attempt)
                print(f"Server Overloaded/Busy. Waiting {wait_time}s... (Attempt {attempt+1}/{max_retries})")
                time.sleep(wait_time)
            else:
                print(f"Gemini API Error: {e}")
                break
    
    return ["-", "-", "-", "-", "United States"]

def get_coordinates_cascading(city, county, state, country, df_lookup):
    """
    Cascading Coordinate Logic:
    1. Sheet2 (US Counties) lookup first (Source of Truth).
    2. Geopy API as fallback if any geo data exists.
    3. Default to '-' if all fail.
    """
    # 1. Sheet2 Lookup (US Only)
    if country == "United States":
        s_val = str(state).strip()
        c_val = str(county).strip()
        match = df_lookup[(df_lookup['State'].astype(str) == s_val) & 
                          (df_lookup['County'].astype(str) == c_val)]
        if not match.empty:
            return f"{match.iloc[0]['Latitude']}, {match.iloc[0]['Longitude']}"

    # 2. Geopy API Fallback (runs if any field is populated)
    geo_parts = [str(p) for p in [city, county, state, country] if p and p != "-"]
    if geo_parts:
        query = ", ".join(geo_parts)
        try:
            time.sleep(1) # Nominatim policy
            location = geolocator.geocode(query, timeout=10)
            if location:
                return f"{location.latitude}, {location.longitude}"
        except:
            pass
            
    return "-"

# --- 2. Load Files ---
df_a = pd.read_excel('Consolidated_Directory.xlsx')
df_b = pd.read_excel('Book_of_Negroes_Copy.xlsx')
df_c = pd.read_excel('US_Counties_Coordinates.xlsx')

df_b.columns = [col.replace(" ", "_") for col in df_b.columns]

# --- 3. Processing ---
print("Extracting data via Gemini (Targeting 15 RPM with Backoff)...")
extracted_geo = df_a['Notes'].apply(lambda x: pd.Series(call_gemini_ner(str(x))))
df_a[['Enslaver', 'Extracted_City', 'Extracted_County', 'Extracted_State', 'Country']] = extracted_geo

print("Calculating Cascading Coordinates...")
df_a['Departure_Coordinates'] = df_a.apply(
    lambda r: get_coordinates_cascading(r['Extracted_City'], r['Extracted_County'], 
                                        r['Extracted_State'], r['Country'], df_c), axis=1)

# --- 4. Cleaning & Renaming ---
if 'Description' in df_a.columns:
    df_a.rename(columns={'Description': 'Origin'}, inplace=True)

if 'Age' in df_a.columns:
    df_a['Birthdate'] = df_a['Age'].apply(
        lambda x: 1783 - int(float(x)) if pd.notna(x) and str(x).replace('.','').isdigit() else "-"
    )
    df_a.drop(columns=['Age'], inplace=True)

df_a['State'], df_a['County'] = "", ""

# --- 5. Merge with Book of Negroes ---
df_a['Match_Name'] = (df_a['First_Name'].fillna('').astype(str).str.strip() + " " + 
                      df_a['Surname'].fillna('').astype(str).str.strip()).str.lower()
df_a['Match_Ship'] = df_a['Ship_Name'].fillna('').astype(str).str.strip().str.lower()

if 'Name' in df_b.columns:
    df_b['Match_Name'] = df_b['Name'].fillna('').astype(str).str.strip().str.lower()
if 'Ship_Name' in df_b.columns:
    df_b['Match_Ship'] = df_b['Ship_Name'].fillna('').astype(str).str.strip().str.lower()

target_cols = ['Ref_Page', 'Origination_Port', 'Departure_Port', 'Departure_Date', 'Arrival_Port_City', 'Ship_Notes']
lookup_cols = ['Match_Ship', 'Match_Name'] + [c for c in target_cols if c in df_b.columns]

df_b_match = df_b[lookup_cols].drop_duplicates(subset=['Match_Ship', 'Match_Name'])
df_final = df_a.merge(df_b_match, on=['Match_Ship', 'Match_Name'], how='left', suffixes=('', '_LKP'))

# --- 6. Reorganize Columns ---
fixed_start = ['Ship_Name', 'Notes', 'Ship_Notes', 'First_Name', 'Surname', 'Birthdate']
geo_block = ['Origin', 'Extracted_City', 'Extracted_County', 'Extracted_State', 'Country', 'State', 'County', 'Departure_Coordinates']

all_cols = df_final.columns.tolist()
remaining = [c for c in all_cols if c not in fixed_start + geo_block and not c.endswith('_LKP') and 'Match_' not in c]
df_final = df_final[fixed_start + remaining + geo_block]

# --- 7. Export & Excel UI ---
output_file = 'Consolidated_Directory_gemini_scraping.xlsx'
df_c = df_c.sort_values(['State', 'County'])

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df_final.to_excel(writer, index=False, sheet_name='Main')
    df_c.to_excel(writer, index=False, sheet_name='Sheet2')

wb = load_workbook(output_file)
ws_main = wb['Main']
header = {cell.value: get_column_letter(cell.column) for cell in ws_main[1]}

# Prepare Dropdowns
unique_states = sorted(df_c['State'].unique().astype(str).tolist())
for i, s in enumerate(unique_states, 1):
    wb['Sheet2'].cell(row=i, column=10).value = s 

dv_state = DataValidation(type="list", formula1=f"Sheet2!$J$1:$J${len(unique_states)}", allow_blank=True)
ws_main.add_data_validation(dv_state)

for r in range(2, ws_main.max_row + 1):
    dv_state.add(ws_main[f"{header['State']}{r}"])
    st_cell = f"{header['State']}{r}"
    formula_co = f'OFFSET(Sheet2!$A$1, MATCH({st_cell}, Sheet2!$B:$B, 0)-1, 0, COUNTIF(Sheet2!$B:$B, {st_cell}), 1)'
    dv_county = DataValidation(type="list", formula1=formula_co, allow_blank=True)
    ws_main.add_data_validation(dv_county)
    dv_county.add(ws_main[f"{header['County']}{r}"])

wb.save(output_file)
print(f"✨ Process complete. Final spreadsheet saved as: {output_file}")