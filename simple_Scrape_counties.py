import pandas as pd
import requests
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io

def scrape_counties_fixed():
    """
    Scrapes all US counties from Wikipedia using a custom User-Agent 
    to bypass the 403 Forbidden error.
    """
    print("🚀 Connecting to Wikipedia...")
    
    url = 'https://en.wikipedia.org/wiki/List_of_United_States_counties_and_county_equivalents'
    
    # Headers to mimic a real browser
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }
    
    try:
        # Step 1: Get the HTML content with headers
        response = requests.get(url, headers=headers)
        response.raise_for_status()  # Check for HTTP errors
        
        # Step 2: Parse the table
        # We use [0] because the main list is the first table on the page
        tables = pd.read_html(io.StringIO(response.text))
        df = tables[0]
        
        print(f"✅ Found {len(df)} counties!")
        return df
        
    except Exception as e:
        print(f"❌ Scraping failed: {e}")
        return None

def format_and_save_excel(df, output_file='US_Counties_Master_List.xlsx'):
    """Format the dataframe and save to Excel with professional styling."""
    if df is None:
        return
    
    print(f"\n📊 Creating formatted Excel file: {output_file}")
    
    # Rename columns based on Wikipedia's table structure
    # Standard columns: County, State, Population, Area, Founded
    df.columns = ['County/Equivalent', 'State', 'Population (2020)', 'Area (sq mi)', 'Founded', 'Footnotes']
    
    # Drop Footnotes and clean strings
    if 'Footnotes' in df.columns:
        df = df.drop('Footnotes', axis=1)
    
    # Clean data: convert to numbers for Excel sorting/math
    for col in ['Population (2020)', 'Area (sq mi)']:
        df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', '').str.split('[').str[0], errors='coerce')

    # Save as CSV for the Streamlit App logic
    df[['County/Equivalent', 'State']].to_csv('us_counties_master.csv', index=False)

    # Save as Formatted Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Counties', index=False)
        
        worksheet = writer.sheets['Counties']
        
        # Define Professional Styles
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Apply Header Style
        for col_num, header in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # Adjust column widths
        widths = {'A': 30, 'B': 10, 'C': 18, 'D': 15, 'E': 20}
        for col, width in widths.items():
            worksheet.column_dimensions[col].width = width
            
        worksheet.freeze_panes = 'A2'

    print(f"✨ Success! File saved as {output_file}")
    print(f"✨ Helper CSV 'us_counties_master.csv' created for Streamlit App.")

if __name__ == '__main__':
    raw_data = scrape_counties_fixed()
    if raw_data is not None:
        format_and_save_excel(raw_data)